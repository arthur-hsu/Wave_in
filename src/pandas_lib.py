import pandas as pd
import openpyxl, os
from openpyxl.styles import Alignment , Font
import src.colorLog as log
import src.config_path as cf


class pandas_Module:
    def __init__(self,excel_path:str='', outside_df = pd.DataFrame()):
        '''
        excel_path : excel file path
        if excel_path exist excel file, load excel file to dataframe
        else load outside_df to dataframe or create empty dataframe.
        '''
        if excel_path == '':
            self.excel_path = os.path.join(cf.get_value('report_dir'),f"{cf.get_value('logname')}.xlsx")
        else:
            self.excel_path = excel_path
        if os.path.exists(self.excel_path):
            _, extension = os.path.splitext(self.excel_path)
            if extension == '.xlsx'  : self.df = pd.read_excel(self.excel_path,engine='openpyxl')
            elif extension == '.csv' : self.df = pd.read_csv(self.excel_path)
        else:
            if not outside_df.empty : self.df = outside_df
            else : self.df = pd.DataFrame(columns=[])

    def write(self, data = dict):
        self.columns = list(data.keys())
        new_data    = pd.DataFrame(data = data, columns = self.columns, index=[0])
        self.df     = pd.concat([self.df, new_data], axis=0, ignore_index=True)
        self.save_to_excel()
        #print(f"\n{self.df[-1:]}\n")

    def creat_df(self):
        log.Logger('[SYSTEM MSG] Create dataframe.' , 'BLUE', 'WHITE',0)
        self.df = pd.DataFrame(columns = self.columns)

    def save_to_excel(self):
        try:
            _, extension = os.path.splitext(self.excel_path)
            columns = self.df.columns.values.tolist()
            if extension == '.xlsx':
                writer = pd.ExcelWriter(path = self.excel_path,engine='openpyxl',mode='w')
                self.df.to_excel(
                excel_writer    = writer,
                sheet_name      = 'RESULT',
                columns         = columns,
                index           = False,
                engine          = 'openpyxl'
                )
                writer.close()
                # self.set_column_width()
            elif extension == '.csv':
                print('extension is %s'%extension)
                self.df.to_csv(self.excel_path, index=False, encoding='utf_8')
            else:
                print('unknow extension')
                return
            # log.Logger('[SYSTEM MSG] Dataframe save to excel.' , 'BLUE', 'WHITE')
        except PermissionError:
            log.color_print('\n\nDo not use microsoft excel to open report file, Use Libreoffice Calc',fore='RED')
            pass

    def set_column_width(self):
        workbook = openpyxl.load_workbook(self.excel_path)
        sheet = workbook.sheetnames
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # 取得欄位的字母表示法
                for cell in col:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                adjusted_width = (max_length + 2) * 1.2  # 計算自動調整後的欄寬
                worksheet.column_dimensions[column].width = adjusted_width  # 設定欄寬

            for row in worksheet.rows:
                for cell in row:
                    # cell.font = font
                    cell.alignment = Alignment(horizontal='center',vertical='center')
        
        workbook.save(self.excel_path)
        workbook.close()

    def __write(self,data=list):
        df_index = 0 if self.df.empty else (int(self.df.index[-1])+1)
        self.df.loc[str(df_index)] = data



